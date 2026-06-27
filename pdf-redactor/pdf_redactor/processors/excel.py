"""Excel processor using openpyxl."""

from __future__ import annotations

import os
import re

from pdf_redactor.detectors.base import Detection

from .base import DocumentProcessor, registry


@registry.register
class ExcelProcessor(DocumentProcessor):
    extensions = ("xlsx", "xlsm")

    def __init__(self, engine, options):
        super().__init__(options)
        self.engine = engine

    def process(self, input_path: str, output_path: str):
        from openpyxl import load_workbook

        from pdf_redactor.engine import RedactionReport

        wb = load_workbook(input_path)
        try:
            # Extract text from every cell to feed detectors.
            text_pages = []
            sheet_names = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                parts = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            parts.append(str(cell.value))
                text_pages.append("\n".join(parts))
                sheet_names.append(sheet_name)

            detections = self._detect(text_pages, sheet_names)
            self._apply(wb, detections)

            os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
            wb.save(output_path)

            report = RedactionReport(pages=len(text_pages), detections=detections)
            report.summary = self._summarize(detections)
            return report
        finally:
            wb.close()

    def _detect(self, text_pages: list[str], sheet_names: list[str]) -> list[Detection]:
        from pdf_redactor.engine import build_detectors

        detectors = build_detectors(self.engine.registry, self.options)
        all_detections: list[Detection] = []
        for detector in detectors:
            for page_num, text in enumerate(text_pages):
                # Reuse PDF detector interface with a dummy document.
                page_detections = detector.detect(_DummyDocument(sheet_names[page_num]), [text])
                all_detections.extend(page_detections)

        if self.options.ml and self.engine.ml_registry is not None:
            for page_num, text in enumerate(text_pages):
                ml_detections = self.engine.ml_registry.detect(
                    _DummyDocument(sheet_names[page_num]), [text], self.options.ml_detectors
                )
                all_detections.extend(ml_detections)

        if self.options.aggressive:
            return all_detections
        return [d for d in all_detections if d.confidence in {"high", "medium"}]

    def _apply(self, wb, detections: list[Detection]) -> None:
        redaction_text = self.options.text or "[REDACTED]"
        # Collect unique sensitive substrings.
        patterns = []
        for d in detections:
            # Escape regex metacharacters in matched text.
            patterns.append(re.escape(d.text))
        if not patterns:
            return
        combined = re.compile("(" + "|".join(patterns) + ")")

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    value = str(cell.value)
                    new_value, count = combined.subn(redaction_text, value)
                    if count:
                        cell.value = new_value

    def _summarize(self, detections: list[Detection]) -> dict[str, dict[str, int]]:
        summary: dict[str, dict[str, int]] = {}
        for d in detections:
            summary.setdefault(d.detector, {"high": 0, "medium": 0, "low": 0})
            summary[d.detector][d.confidence] += 1
        return summary


class _DummyDocument:
    """Minimal stand-in for fitz.Document so text detectors can call search_for."""

    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name

    def load_page(self, page_num: int):
        return _DummyPage(self.sheet_name)

    def __len__(self):
        return 1


class _DummyPage:
    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name

    def search_for(self, text: str):
        # Excel redaction does not use bounding boxes.
        return []
